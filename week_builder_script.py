import sqlite3
import random
from rich import print
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side


# def get_lecture_id_by_code(code):
#     lecture_id = None
#     for row2 in lectures:
#         if row2[1] == code:
#             lecture_id = row2[0]
#             break
    
#     return lecture_id

def is_professor_available(professor_id, hour, day, professors_time):
    for row in professors_time:
        if row[1] == professor_id and row[2] == hour and row[3] == day:
            return True
    return False

# def get_professor_conflicts_num(lectures, hour, day):
#     conflicts_num = 0
#     professors_ids = []

#     for lec in lectures:
#         lec_id = get_lecture_id_by_code(lec[0])
                
#         for row2 in lectures_professores:
#             if row2[2] == lec_id:
#                 if row2[1] in professors_ids:
#                     conflicts_num += 1
#                 else:
#                     if not is_professor_available(row2[1], hour, day):
#                         conflicts_num += 1
#                     professors_ids.append(row2[1])
    
#     return conflicts_num

# def get_student_conflicts_num(lectures):
#     conflicts_num = 0
#     students_ids = []

#     for lec in lectures:
#         lec_id = get_lecture_id_by_code(lec[0])
                
#         for row2 in lectures_students:
#             if row2[2] == lec_id:
#                 if row2[1] in students_ids:
#                     conflicts_num += 1
#                 else:
#                     students_ids.append(row2[1])
    
#     return conflicts_num

def get_sequence_subset(list_, subset_size):
    return [
        list_[i: i + subset_size]
        for i in range(len(list_) - subset_size + 1)
    ]

def get_items_difference_sum(list_):
    result = 0
    for i in range(1, len(list_)):
        result += abs(list_[i] - list_[i-1])
    return result

def get_items_distance_from_point_sum(list_, point):
    result = 0
    for item in list_:
        result += abs(item - point)
    return result

def get_shared_items_between_lists(list1, list2):
    shared_items = []

    for item1 in list1:
        for item2 in list2:
            if item1 == item2:
                shared_items.append(item1)
                break
    
    return shared_items

def get_lecture_day_available_hours(lecture, week, day, professors_time):
    empty_hours = []
    for hour in week[day]:
        if not is_professor_available(lecture["professorId"], hour, day, professors_time):
            continue

        for lec_ in week[day][hour]:
            # there can only be one professor for a lecture hour
            if lec_["professorId"] == lecture["professorId"]:
                break
            # we check for conflicts only for lectures with the same year
            if lec_["year"] == lecture["year"] and get_shared_items_between_lists(
                lec_["studentIds"],
                lecture["studentIds"]
            ):
                break
        else:
            empty_hours.append(hour)
    
    return empty_hours

def score_hours(hours):
    diff = get_items_difference_sum(hours) - len(hours) + 1
    # distance_from_point = 0#get_items_distance_from_point_sum(hours, 12) / (len(hours) * 2)

    # return - diff - distance_from_point
    return - diff

def get_best_hours(hours_list, acceptable_threshold = None):
    max_score = None
    max_score_hours = None

    for hours in hours_list:
        score = score_hours(hours)

        if acceptable_threshold is not None and score >= acceptable_threshold:
            return hours, score
        
        if max_score is None or score > max_score:
            max_score = score
            max_score_hours = hours
    
    return max_score_hours, max_score

def place_lecture_in_week(lecture, week, professors_time):
    best_hours, best_score, best_day = None, None, None
    for day in week:
        empty_hours = get_lecture_day_available_hours(lecture, week, day, professors_time)

        hours_list = get_sequence_subset(empty_hours, lecture["hours"])

        hours, score = get_best_hours(hours_list)

        if score is None:
            continue

        if best_score is None or score > best_score:
            best_score = score
            best_hours = hours
            best_day = day

    if best_hours is None:
        if lecture["hours"] == 1:
            raise Exception("Could not place lecture " + lecture["code"])
            
        splitted_lectures = split_lecture_time(lecture)

        for splitted_lecture in splitted_lectures:
            week = place_lecture_in_week(splitted_lecture, week, professors_time)
        
        return week

    for hour in best_hours:
        week[best_day][hour].append(lecture.copy())
    
    return week

def split_lecture_time(lecture):
    hours = lecture["hours"]
    lec1 = lecture.copy()
    lec1["hours"] = hours // 2
    lec2 = lecture.copy()
    lec2["hours"] = hours - lec1["hours"]

    return [lec1, lec2]

def build_week(is_random = False):
    conn = sqlite3.connect("Databases/UniversityDb.db")
    cursor = conn.cursor()

    lectures = cursor.execute("SELECT * FROM LectureTb").fetchall()
    lectures_professores = cursor.execute("SELECT * FROM LectureProfessorTb").fetchall()
    lectures_students = cursor.execute("SELECT * FROM LectureStudentTb").fetchall()
    professors_time = cursor.execute("SELECT * FROM TimeProfessorTb").fetchall()
    students = cursor.execute("SELECT * FROM StudentTb").fetchall()
    professors = cursor.execute("SELECT * FROM ProfessorTb").fetchall()

    WEEK = dict(
        (day, 
            dict([
                (hour, [])
                for hour in [
                    8, 9, 10, 11, 13, 14, 15, 16, 17
                ]
            ])
        )
        for day in [
            "Monday",
            "Tuesday",
            "Wednesday",
            "Thursday",
            "Friday",
        ]
    )

    LECTURES = [
        {
            'code': row[1],
            "professorId": row2[1],
            "studentIds": [
                row3[1]
                for row3 in lectures_students
                if row2[0] == row3[2]
            ],
            "hours": row[3],
            "year": row[4],
        }
        for row in lectures
        for row2 in lectures_professores
        if row[0] == row2[2]
    ]

    if is_random:
        random.shuffle(LECTURES)
    else:
        LECTURES.sort(key=lambda x: -x["hours"])

    for lec in LECTURES:
        if not lec["studentIds"]:
            print(f"Lecture: {lec["code"]}, for Professor: {lec["professorId"]} has been dropped!")
            print("Cause: No students take this lecture.")
            # continue

        # lec = get_fully_detailed_lecture(lec)
        week = place_lecture_in_week(lec, WEEK, professors_time)

    for day in week:
        for hour in week[day]:
            for l, lec in enumerate(week[day][hour]):
                week[day][hour][l] = get_fully_detailed_lecture(lec, lectures, students, professors)

    return week

def get_fully_detailed_lecture(lecture, lectures, students, professors):
    for row in lectures:
        if lecture["code"] == row[1]:
            lecture["name"] = row[2]
            break

    for row in professors:
        if lecture["professorId"] == row[0]:
            lecture["professorName"] = row[2]
            del lecture["professorId"] # it is safer to remove any database info
            lecture["professorNumber"] = row[1]
            break
    
    lecture["studentNumbers"] = []
    for student_id in lecture["studentIds"]:
        for row in students:
            if student_id == row[0]:
                lecture["studentNumbers"].append(row[1])
                break
    
    del lecture["studentIds"]
    
    return lecture

def combine_sequenced_lectures(week):
    for day in week:
        for hour in week[day]:
            for l, lec in enumerate(week[day][hour]):
                week[day][hour][l]["count"] = 1

    for day in week:
        reversed_hours = list(week[day].keys())[::-1]
        for h, hour in enumerate(reversed_hours[:-1]):
            to_delete = []
            for l in range(len(week[day][hour])-1, -1, -1):
                lec = week[day][hour][l]
            # for l, lec in enumerate(week[day][hour]):
                for l2, lec2 in enumerate(week[day][reversed_hours[h+1]]):
                    if lec["code"] == lec2["code"] and lec["professorNumber"] == lec2["professorNumber"]:
                        week[day][reversed_hours[h+1]][l2]["count"] += week[day][hour][l]["count"]
                        week[day][hour].pop(l)
                        # to_delete.append(l)

                        break
    
    return week

def tableize_combined_week_by_year(combined_week, remove_unnecessary_nones = False):
    cols_availibility_counter = {}
    table = {
        "rows": {},
        "cols": {},
    }
    global_row_pointer = 0

    for day in combined_week:
        table["rows"][day] = {}

        for hour in combined_week[day]:
            table["rows"][day][hour] = {}

            if not remove_unnecessary_nones:
                for year in cols_availibility_counter:
                    for i, item in enumerate(cols_availibility_counter[year]):
                        if item != 0:
                            table["cols"][year][i].append(None)

            lectures = combined_week[day][hour][::-1]

            for l in range(len(lectures)-1, -1, -1):
                lec = lectures[l]

                if lec["year"] not in cols_availibility_counter:
                    cols_availibility_counter[lec["year"]] = []
                    table["cols"][lec["year"]] = []

                for year in cols_availibility_counter:

                    if lec["year"] != year:
                        continue

                    for i, item in enumerate(cols_availibility_counter[year]):
                        if item == 0:
                            table["cols"][year][i].append(lec)
                            cols_availibility_counter[year][i] = lec["count"]
                            lectures.pop(l)
                            break
                    else:
                        table["cols"][year].append([{} for _ in range(global_row_pointer)] + [lec])
                        cols_availibility_counter[year].append(lec["count"])


            for year in cols_availibility_counter:
                for i, item in enumerate(cols_availibility_counter[year]):
                    if item == 0:
                        table["cols"][year][i].append({})
                    else:
                        cols_availibility_counter[year][i] -= 1

            global_row_pointer += 1

    table["cols"] = dict(sorted(table["cols"].items())) # to sort keys

    return table

def build_week_html_content(tableized_week):
    style = """<style>
    table, th, tr {
        border: 1px solid black;
    }
    td, th {
        padding: 5px;
    }
    .bordered {
        border: 1px solid black;
        text-align: center;
        vertical-align: middle;
    }
    .vertical_writing {
        transform: rotate(-90deg);
    }
    table {
        border-collapse: collapse;
    }
    .header_row {
        /* background-color: rgb(255, 89, 89); */
        background-color: aqua;
    }
    .lec_cell {
        background-color: rgb(224, 224, 224);
    }
    </style>"""

    table_content = style + "<table>"

    header_content = """<tr>
    <td class="bordered header_row">Day</td>
    <td class="bordered header_row">Time</td>""" + "".join(
        f"""
        <td class="bordered header_row" colspan="{len(tableized_week["cols"][year])}">Year {year}</td>
        """
        for year in tableized_week["cols"]
    ) + """
</tr>"""
    

    global_table_row_pointer = 0

    for day in tableized_week["rows"]:
        table_content += header_content
        
        days_content = f"""<td class="bordered vertical_writing" rowspan="{len(tableized_week["rows"][day])+1}">{day}</td>"""
        
        rows_content = []
        for hour in tableized_week["rows"][day]:
            rows_content.append(f"""<tr><td class="bordered">{hour}:00 ~ {hour}:50</td>""")
        
        for year in tableized_week["cols"]:
            for col in tableized_week["cols"][year]:
                for l, lec in enumerate(col[global_table_row_pointer : global_table_row_pointer + len(rows_content)]):
                    if lec is None:
                        continue

                    if lec == {}:
                        rows_content[l] += "<td></td>"
                        continue

                    span = lec["count"]

                    rows_content[l] += f"""<td rowspan="{span}" class="bordered lec_cell">{lec["name"]}<br>{lec["professorName"]}</td>"""
        
        global_table_row_pointer += len(rows_content)

        for i in range(len(rows_content)):
            rows_content[i] += "</tr>"
        
        days_content += "".join(rows_content)

        table_content += days_content

    table_content += "</table>"

    return table_content

def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get column letter (A, B, C, etc.)
        
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        ws.column_dimensions[col_letter].width = max_length

def auto_adjust_row_height(ws):
    for row in ws.iter_rows():
        max_height = 15  # Default row height
        for cell in row:
            if cell.value:
                lines = str(cell.value).count("\n") + 1  # Count line breaks
                max_height = max(max_height, lines * 15)  # Adjust for multiline text
        
        ws.row_dimensions[row[0].row].height = max_height

def get_max_dimentions(tableized_week):
    rows_num = 0
    for day in tableized_week["rows"]:
        rows_num += len(tableized_week["rows"][day]) + 1 # fore headers

    cols_num = 2
    for year in tableized_week["cols"]:
        years_columns_num = len(tableized_week["cols"][year])
        cols_num += years_columns_num
    
    return rows_num, cols_num

def build_week_excel_file(tableized_week):
    header_fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")
    cell_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")  # Center alignment
    day_alignment = Alignment(textRotation=90, horizontal="center", vertical="center")  # Center alignment
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    wb = openpyxl.Workbook()
    ws = wb.active

    def add_header(ws, row, col):
        cell = ws.cell(row=1+row, column=1+col, value=f"Day")
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = border

        cell = ws.cell(row=1+row, column=1+col+1, value=f"Time")
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = border

        col += 2

        for year in tableized_week["cols"]:
            years_columns_num = len(tableized_week["cols"][year])
            
            cell = ws.cell(row=1+row, column=1+col, value=f"Year {year}")
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

            ws.merge_cells(
                start_row=1+row,
                start_column=1+col,
                end_row=1+row,
                end_column=1+col+years_columns_num -1
            )

            col += years_columns_num

    global_row_pointer = 0
    global_table_row_pointer = 0
    global_col_pointer = 0


    for day in tableized_week["rows"]:
        add_header(ws, global_row_pointer, global_col_pointer)
        global_row_pointer += 1

        cell = ws.cell(row=1+global_row_pointer, column=1+global_col_pointer, value=day)
        # cell.fill = cell_fill
        cell.alignment = day_alignment
        cell.border = border

        ws.merge_cells(
            start_row=1+global_row_pointer,
            start_column=1+global_col_pointer,
            end_row=1+global_row_pointer+len(tableized_week["rows"][day]) -1,
            end_column=1+global_col_pointer
        )

        local_row_pointer = global_row_pointer
        for hour in tableized_week["rows"][day]:
            cell = ws.cell(row=1+local_row_pointer, column=1+global_col_pointer+1, value=f"{hour}:00 ~ {hour}:50")
            # cell.fill = cell_fill
            cell.alignment = alignment
            cell.border = border
            local_row_pointer += 1
        
        hours_num = len(tableized_week["rows"][day])

        local_col_pointer = global_col_pointer
        for year in tableized_week["cols"]:
            # to_skip = 0
            for col in tableized_week["cols"][year]:
                local_row_pointer = global_row_pointer
                for lec in col[global_table_row_pointer : global_table_row_pointer + hours_num]:
                    local_row_pointer += 1


                    # if to_skip > 0:
                    #     to_skip -= 1
                    #     continue

                    if lec is None:
                        continue

                    if lec == {}:
                        continue

                    to_skip = lec["count"] -1
                    
                    cell = ws.cell(
                        row=local_row_pointer,
                        column=1+local_col_pointer+2,
                        value=lec["name"] + "\n" + lec["professorName"]
                    )
                    cell.fill = cell_fill
                    cell.alignment = alignment
                    cell.border = border
                    ws.merge_cells(
                        start_row=local_row_pointer,
                        start_column=1+local_col_pointer+2,
                        end_row=local_row_pointer+to_skip,
                        end_column=1+local_col_pointer+2
                    )

                    # local_row_pointer += to_skip
                
                local_col_pointer += 1
        
        global_row_pointer += hours_num
        global_table_row_pointer += hours_num

    rows_num, cols_num = get_max_dimentions(tableized_week)

    for i in range(rows_num):
        cell = ws.cell(row=1+i, column=cols_num)
        b = cell.border
        cell.border = Border(left=b.left, top=b.top, right=Side(style="thin"), bottom=b.bottom)

    for i in range(cols_num):
        cell = ws.cell(row=rows_num, column=1+i)
        b = cell.border
        cell.border = Border(left=b.left, top=b.top, right=b.right, bottom=Side(style="thin"))

    auto_adjust_column_width(ws)
    # auto_adjust_row_height(ws)

    # wb.save("merged_cells.xlsx")
    return wb